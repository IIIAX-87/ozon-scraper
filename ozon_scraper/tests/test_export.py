"""Tests for single-sheet image export."""
from __future__ import annotations

import io
import os
import tempfile
import unittest
from unittest.mock import MagicMock, patch

import openpyxl
import requests
from openpyxl.worksheet.worksheet import Worksheet

from export_xlsx import export_api_products


class FakeImageResponse:
    def __init__(self):
        # Minimal valid 1x1 red PNG
        self.content = (
            b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
            b"\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDATx\x9cc\xf8\xcf\xc0\x00\x00"
            b"\x03\x01\x01\x00\xc9\xfe\x92\xef\x00\x00\x00\x00IEND\xaeB`\x82"
        )
        self.status_code = 200

    def raise_for_status(self):
        pass


class TestExportSingleSheet(unittest.TestCase):
    def _make_product(self, image_url: str = "https://cdn.ozone.ru/1001.jpg"):
        product = MagicMock()
        product.product_id = 1001
        product.offer_id = "SKU-001"
        product.sku = None
        product.name = "Product 1"
        product.barcode = "BAR-1"
        product.category_id = 10
        product.description_category_id = 11
        product.type_id = 20
        product.description = "Test description"
        product.price = 1000.0
        product.old_price = 1200.0
        product.retail_price = 900.0
        product.marketing_price = 950.0
        product.min_price = 800.0
        product.currency_code = "RUB"
        product.vat = 20.0
        product.status = ""
        product.visibility = ""
        product.archived = False
        product.discounted = False
        product.height = 10
        product.depth = 20
        product.width = 30
        product.dimension_unit = "mm"
        product.weight = 100.0
        product.weight_unit = "g"
        product.volume_weight = 120.0
        product.primary_image = image_url
        product.images = [image_url]
        product.attributes = [
            {
                "attribute_id": 1,
                "complex_id": 0,
                "values": [{"dictionary_value_id": 0, "value": "Red"}],
            }
        ]
        product.complex_attributes = []
        return product

    def test_export_inlines_images(self):
        product = self._make_product()

        with patch.object(requests.Session, "get", return_value=FakeImageResponse()):
            with tempfile.TemporaryDirectory() as tmpdir:
                path = os.path.join(tmpdir, "test.xlsx")
                result = export_api_products([product], path)
                self.assertTrue(os.path.exists(result))

                wb = openpyxl.load_workbook(result)
                ws = wb["Товары"]
                photo_col = None
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=1, column=col).value == "Фото":
                        photo_col = col
                        break
                self.assertIsNotNone(photo_col)
                assert photo_col is not None
                photo_cell = ws.cell(row=2, column=photo_col)
                self.assertIsNone(photo_cell.comment)
                self.assertTrue(len(ws._images) > 0)  # type: ignore[attr-defined]

    def test_export_fallback_link_when_download_fails(self):
        product = self._make_product(image_url="https://cdn.ozone.ru/missing.jpg")

        with patch.object(requests.Session, "get", side_effect=requests.RequestException("boom")):
            with tempfile.TemporaryDirectory() as tmpdir:
                path = os.path.join(tmpdir, "test.xlsx")
                result = export_api_products([product], path)
                self.assertTrue(os.path.exists(result))

                wb = openpyxl.load_workbook(result)
                ws = wb["Товары"]
                photo_col = None
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=1, column=col).value == "Фото":
                        photo_col = col
                        break
                self.assertIsNotNone(photo_col)
                assert photo_col is not None
                photo_cell = ws.cell(row=2, column=photo_col)
                self.assertIsNone(photo_cell.comment)
                self.assertEqual(photo_cell.value, "https://cdn.ozone.ru/missing.jpg")


if __name__ == "__main__":
    unittest.main()
