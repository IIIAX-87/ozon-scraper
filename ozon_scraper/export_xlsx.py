"""Export API-collected Ozon products to a single-sheet .xlsx file with images."""
from __future__ import annotations

import io
import os
import re
from typing import Any, Dict, List, Optional, cast

import openpyxl
import requests
from openpyxl.cell.cell import Cell
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

try:
    from PIL import Image as PILImage  # type: ignore[import-not-found]
except Exception:  # pragma: no cover
    PILImage = None  # type: ignore[misc,assignment]


def _set_header_row(ws: Worksheet, headers: List[str]) -> None:
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font


def _format_attributes(
    attributes: List[Dict[str, Any]],
    complex_attributes: List[Dict[str, Any]],
) -> str:
    """Convert attributes to a readable string."""
    parts: List[str] = []

    for attr in attributes or []:
        values = [str(v.get("value", "")) for v in attr.get("values", []) or []]
        if values:
            parts.append(f"{attr.get('attribute_id')}: {', '.join(values)}")

    for cidx, complex_attr in enumerate(complex_attributes or []):
        for attr in complex_attr.get("attributes", []) or []:
            values = [str(v.get("value", "")) for v in attr.get("values", []) or []]
            if values:
                parts.append(
                    f"complex[{cidx}].{attr.get('attribute_id')}: {', '.join(values)}"
                )

    return "\n".join(parts)


def _sanitize_filename(name: str) -> str:
    return re.sub(r"[^\w\-_. ]", "_", name).strip() or "image"


def _download_image(
    url: str,
    session: requests.Session,
    timeout: int = 20,
    max_retries: int = 2,
) -> Optional[bytes]:
    if not url:
        return None

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
        "Referer": "https://www.ozon.ru/",
    }

    for attempt in range(max_retries + 1):
        try:
            resp = session.get(url, headers=headers, timeout=timeout)
            resp.raise_for_status()
            content = resp.content
            if content:
                return content
        except Exception:
            if attempt == max_retries:
                return None
    return None


def _resize_image_bytes(data: bytes, max_width: int = 200, max_height: int = 200) -> bytes:
    """Resize image keeping aspect ratio, return as PNG bytes."""
    if PILImage is None:
        raise RuntimeError("Pillow is required for image processing")

    img = PILImage.open(io.BytesIO(data))

    # Convert palette/transparent images to RGBA before saving as PNG.
    if img.mode not in ("RGB", "RGBA"):
        img = img.convert("RGBA") if img.mode == "P" and "transparency" in img.info else img.convert("RGB")

    img.thumbnail((max_width, max_height))
    out = io.BytesIO()
    img.save(out, format="PNG")
    return out.getvalue()


def export_api_products(
    products: List[Any],
    output_path: Optional[str] = None,
    errors: Optional[List[str]] = None,
    image_column: str = "Фото",
    image_max_width: int = 200,
    image_max_height: int = 200,
) -> str:
    """Export products to a single worksheet with inline images and all attributes."""
    output_path = output_path or os.getenv("OUTPUT_FILE", "ozon_products_api.xlsx")

    headers = [
        "product_id",
        "offer_id",
        "sku",
        "name",
        "barcode",
        "category_id",
        "description_category_id",
        "type_id",
        "description",
        "price",
        "old_price",
        "retail_price",
        "marketing_price",
        "min_price",
        "currency_code",
        "vat",
        "status",
        "visibility",
        "archived",
        "discounted",
        "height",
        "depth",
        "width",
        "dimension_unit",
        "weight",
        "weight_unit",
        "volume_weight",
        image_column,
        "attributes",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to create workbook sheet")
    ws.title = "Товары"

    _set_header_row(ws, headers)

    photo_col_index = headers.index(image_column) + 1
    row_height_pt = int(image_max_height * 0.75) + 6

    session = requests.Session()

    for row, product in enumerate(products, 2):
        ws.cell(row=row, column=1, value=product.product_id)
        ws.cell(row=row, column=2, value=product.offer_id)
        ws.cell(row=row, column=3, value=product.sku)
        ws.cell(row=row, column=4, value=product.name)
        ws.cell(row=row, column=5, value=product.barcode)
        ws.cell(row=row, column=6, value=product.category_id)
        ws.cell(row=row, column=7, value=product.description_category_id)
        ws.cell(row=row, column=8, value=product.type_id)
        ws.cell(row=row, column=9, value=product.description)
        ws.cell(row=row, column=10, value=product.price)
        ws.cell(row=row, column=11, value=product.old_price)
        ws.cell(row=row, column=12, value=product.retail_price)
        ws.cell(row=row, column=13, value=product.marketing_price)
        ws.cell(row=row, column=14, value=product.min_price)
        ws.cell(row=row, column=15, value=product.currency_code)
        ws.cell(row=row, column=16, value=product.vat)
        ws.cell(row=row, column=17, value=product.status)
        ws.cell(row=row, column=18, value=product.visibility)
        ws.cell(row=row, column=19, value=product.archived)
        ws.cell(row=row, column=20, value=product.discounted)
        ws.cell(row=row, column=21, value=product.height)
        ws.cell(row=row, column=22, value=product.depth)
        ws.cell(row=row, column=23, value=product.width)
        ws.cell(row=row, column=24, value=product.dimension_unit)
        ws.cell(row=row, column=25, value=product.weight)
        ws.cell(row=row, column=26, value=product.weight_unit)
        ws.cell(row=row, column=27, value=product.volume_weight)

        # Attributes column
        attr_text = _format_attributes(product.attributes, product.complex_attributes)
        ws.cell(row=row, column=29, value=attr_text)

        # Image column (keep empty; image anchored later)
        ws.cell(row=row, column=photo_col_index, value="")
        ws.row_dimensions[row].height = row_height_pt

    # Download and insert images
    for row, product in enumerate(products, 2):
        image_url = product.primary_image or (product.images[0] if product.images else "")
        image_data = _download_image(image_url, session)
        cell = cast(Cell, ws.cell(row=row, column=photo_col_index))
        if image_data:
            try:
                resized = _resize_image_bytes(image_data, image_max_width, image_max_height)
                img = OpenpyxlImage(io.BytesIO(resized))
                img.width = image_max_width
                img.height = image_max_height
                ws.add_image(img, cell.coordinate)
                cell.value = ""
            except Exception:
                # If image processing fails, keep the URL as a clickable link in the cell.
                cell.hyperlink = image_url  # type: ignore[assignment]
                cell.value = image_url
        elif image_url:
            # Keep a clickable link if download failed.
            cell.hyperlink = image_url  # type: ignore[assignment]
            cell.value = image_url

    # Adjust column widths
    ws.column_dimensions[get_column_letter(photo_col_index)].width = int(image_max_width / 6)
    for col in range(1, len(headers) + 1):
        if col == photo_col_index:
            continue
        max_length = len(headers[col - 1])
        for row in range(2, len(products) + 2):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 60)

    # Optional errors sheet
    if errors:
        ws_errors = wb.create_sheet(title="Ошибки")
        error_headers = ["#", "message"]
        _set_header_row(ws_errors, error_headers)
        for row, err in enumerate(errors, 2):
            ws_errors.cell(row=row, column=1, value=row - 1)
            ws_errors.cell(row=row, column=2, value=err)
        for col in range(1, 3):
            max_length = len(error_headers[col - 1])
            for row in range(2, len(errors) + 2):
                value = ws_errors.cell(row=row, column=col).value
                if value is not None:
                    max_length = max(max_length, len(str(value)))
            ws_errors.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 80)

    wb.save(output_path)
    session.close()
    return os.path.abspath(output_path)


def export_products(
    products: List[Any],
    output_path: Optional[str] = None,
) -> str:
    """Legacy compatibility for web-scraper products."""
    output_path = output_path or os.getenv("OUTPUT_FILE", "ozon_products.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to create workbook sheet")
    ws.title = "Товары"

    headers = [
        "Название",
        "Цена",
        "Старая цена",
        "Скидка",
        "Рейтинг",
        "Отзывы",
        "Ссылка",
        "Изображение",
    ]

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row, product in enumerate(products, 2):
        ws.cell(row=row, column=1, value=product.name)
        ws.cell(row=row, column=2, value=product.price)
        ws.cell(row=row, column=3, value=product.old_price)
        ws.cell(row=row, column=4, value=product.discount)
        ws.cell(row=row, column=5, value=product.rating)
        ws.cell(row=row, column=6, value=product.reviews)
        ws.cell(row=row, column=7, value=product.url)
        ws.cell(row=row, column=8, value=product.image)

    for col in range(1, len(headers) + 1):
        max_length = len(headers[col - 1])
        for row in range(2, len(products) + 2):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 80)

    wb.save(output_path)
    return os.path.abspath(output_path)
